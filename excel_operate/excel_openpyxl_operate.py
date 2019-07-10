# -*- coding: utf-8 -*- 
# @Time : 2019/7/10 20:11 
# @Author : FengHao 
# @Site :  
# @File : excel_openpyxl_operate 
# @Software: PyCharm Community Edition

import json
from openpyxl import load_workbook, Workbook

from common import log_learn
FIELD_TYPE_INT = 'field_type_int'
FIELD_TYPE_STR = 'field_type_str'
FIELD_NAME = 'field_name'
FIELD_TYPE = 'field_type'

class CPyXLReadExcel:
    def __init__(self, logger, excel_file, excel_sheet, field_info):
        self.name = 'CPyXLReadExcel'
        self.logger = logger
        self.excel_file = excel_file
        self.excel_sheet = excel_sheet
        self.field_info = field_info

    def read_excel_file(self, excel_data):
        """
        读取excel file中数据
        :param excel_data:
        :return:
        """
        self.logger.info('file: %s, sheet: %s' % (self.excel_file, self.excel_sheet))
        wb = load_workbook(self.excel_file)

        wb.active = wb.get_sheet_by_name(self.excel_sheet)
        ws = wb.active

        for row in ws.iter_rows(min_row=2):
            t_data = dict()
            # print 'row:', row
            # print self.field_info
            for index, field_info in enumerate(self.field_info):
                # print 'index: %s %d, %s' % (key, index, row[index].value)
                field_name = field_info[FIELD_NAME]
                field_type = field_info[FIELD_TYPE]
                t_value = row[index].value if row[index].value is not None else ''
                if field_type == FIELD_TYPE_INT:
                    if not isinstance(t_value, int):
                        t_value = int(t_value)
                if field_type == FIELD_TYPE_STR:
                    # if not isinstance(t_value, unicode) and not isinstance(t_value, str): --python2
                    if not isinstance(t_value, bytes) and not isinstance(t_value, str):
                        # print '%s' % t_value,
                        # print type(t_value)
                        t_value = str(t_value)
                t_data[field_name] = t_value
            # print t_data
            self.logger.debug(json.dumps(t_data))
            excel_data.append(t_data)


class CPyXLWriteExcel:
    def __init__(self, logger, excel_write_file, sheet_name, field_info):
        self.name = 'CPyXLWriteExcel'
        self.logger = logger
        self.excel_write_file = excel_write_file
        self.sheet_name = sheet_name
        self.field_info = field_info

    def write_excel_file(self, excel_data):
        """
        write data to excel file
        :param excel_data:
        :return:
        """
        self.logger.info('file: %s, sheet: %s' % (self.excel_write_file, self.sheet_name))
        wb = Workbook()
        ws = wb.active
        ws.title = self.sheet_name

        table_title = self.field_info
        for col, value in enumerate(table_title):
            c = col + 1
            ws.cell(row=1, column=c).value = table_title[col]

        for row in excel_data:
            row_value = list()
            for item in self.field_info:
                row_value.append(row[item] if item in row else '')
            # print row_value
            ws.append(row_value)

        wb.save(filename=self.excel_write_file)

if __name__ == '__main__':
    g_log_name = "logs/excel_openpyxl_operate.log"
    g_log_info = 'excel_openpyxl_operate'
    g_logger = log_learn.LogHelper(logfile=g_log_name, logger=g_log_info).getlog()
    g_logger.info('begin')
    g_excel_file = './data/finance_industry_tag_map.xlsx'
    g_sheet_name = 'finance_industry'
    g_excel_field_info = [{FIELD_NAME: 'title', FIELD_TYPE: FIELD_TYPE_STR},
                          {FIELD_NAME: 'tid', FIELD_TYPE: FIELD_TYPE_INT},
                          {FIELD_NAME: 'cate_1', FIELD_TYPE: FIELD_TYPE_STR},
                          {FIELD_NAME: 'cate_2', FIELD_TYPE: FIELD_TYPE_STR},
                          {FIELD_NAME: 'fcate_1', FIELD_TYPE: FIELD_TYPE_STR},
                          {FIELD_NAME: 'fcate_2', FIELD_TYPE: FIELD_TYPE_STR}]
    excel_read = CPyXLReadExcel(g_logger, g_excel_file, g_sheet_name, g_excel_field_info)
    g_excel_data = list()
    excel_read.read_excel_file(g_excel_data)
    print(g_excel_data)
    g_logger.info('end')
